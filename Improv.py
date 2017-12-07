# -*- coding: utf-8 -*-
"""
Created on Mon Oct 30 12:13:43 2017

@author: Julien Arbellini
""" 

import os
from openpyxl import load_workbook
import numpy as np

#Menu de base
os.system('cls')
Menu = input("Menu\n1 - Apprends des accords\n2 - Assistant impro\n")
while Menu != "1" and "2":
    os.system('cls')
    Menu = input("Appuie sur 1 ou 2\nMenu\n1 - Apprends des accords\n2 - Assistant impro\n")


wb = load_workbook('Les accords.xlsx')
sheet = wb.get_sheet_by_name('Feuille 1')
Niveau = input("Niveau ?\n")
Style = input("Style ?\n")
TNiv = len(sheet['C'])
Tab1 = []

for i in range(1,TNiv):
    ValNiv = sheet.cell(row = i, column=3).value
    ValSty = sheet.cell(row = i, column=1).value
    ValAcc = sheet.cell(row = i, column = 2).value
    if ValNiv == Niveau and ValSty == Style:
        Tab1.append(ValAcc)
al = np.random.randint(0,len(Tab1))

print(Tab1[al])

